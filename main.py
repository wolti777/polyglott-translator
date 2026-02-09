import random
from datetime import timedelta, datetime
from io import BytesIO
from typing import Optional

from fastapi import FastAPI, Request, Depends, Form, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session
from openpyxl import Workbook
from pydantic import BaseModel

from database import engine, get_db, Base, DB_PATH
from models import User, Glossary, GlossaryEntry, UserApiKey
from auth import (
    create_access_token,
    create_user,
    authenticate_user,
    get_user_by_username,
    get_user_by_email,
    get_current_user,
    get_password_hash,
    verify_password,
    encrypt_api_key,
    decrypt_api_key,
    create_verification_token,
    verify_email_token,
    create_password_reset_token,
    verify_password_reset_token,
    send_verification_email,
    send_password_reset_email,
    ACCESS_TOKEN_EXPIRE_MINUTES,
)
from translator import translate_to_all_languages, get_trial_days_remaining

# Create database tables
Base.metadata.create_all(bind=engine)

# Migrate: add new language columns if they don't exist
def migrate_add_language_columns():
    import sqlite3
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(glossary_entries)")
    existing_cols = {row[1] for row in cursor.fetchall()}
    for col in ["french", "italian", "portuguese", "dutch", "russian"]:
        if col not in existing_cols:
            cursor.execute(f"ALTER TABLE glossary_entries ADD COLUMN {col} VARCHAR(500)")
    if "learning_rate" not in existing_cols:
        cursor.execute("ALTER TABLE glossary_entries ADD COLUMN learning_rate INTEGER DEFAULT 0")
    if "total_learning_rate" not in existing_cols:
        cursor.execute("ALTER TABLE glossary_entries ADD COLUMN total_learning_rate INTEGER DEFAULT 0")
    # Ensure existing NULL values are set to 0
    cursor.execute("UPDATE glossary_entries SET learning_rate = 0 WHERE learning_rate IS NULL")
    cursor.execute("UPDATE glossary_entries SET total_learning_rate = 0 WHERE total_learning_rate IS NULL")

    # Migrate users table: add email, email_verified, created_at
    cursor.execute("PRAGMA table_info(users)")
    user_cols = {row[1] for row in cursor.fetchall()}
    if "email" not in user_cols:
        cursor.execute("ALTER TABLE users ADD COLUMN email VARCHAR(255)")
    if "email_verified" not in user_cols:
        cursor.execute("ALTER TABLE users ADD COLUMN email_verified BOOLEAN DEFAULT 0")
    if "created_at" not in user_cols:
        cursor.execute("ALTER TABLE users ADD COLUMN created_at DATETIME")
        # Set created_at for existing users
        cursor.execute("UPDATE users SET created_at = datetime('now') WHERE created_at IS NULL")

    if "language_config" not in user_cols:
        cursor.execute("ALTER TABLE users ADD COLUMN language_config TEXT")

    # Auto-verify first user (admin)
    cursor.execute("UPDATE users SET email_verified = 1 WHERE id = 1")

    conn.commit()
    conn.close()

migrate_add_language_columns()

app = FastAPI(title="Polyglot Translator")

# Mount static files and templates
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


def get_base_url(request: Request) -> str:
    """Get the base URL for email links."""
    # Use X-Forwarded headers if behind a proxy
    proto = request.headers.get("x-forwarded-proto", request.url.scheme)
    host = request.headers.get("x-forwarded-host", request.headers.get("host", "localhost"))
    return f"{proto}://{host}"


class TranslateRequest(BaseModel):
    text: str
    source_language: Optional[str] = None
    target_languages: Optional[list] = None
    enabled_services: Optional[dict] = None
    explanation_services: Optional[dict] = None


class SaveGlossaryRequest(BaseModel):
    spanish: Optional[str] = None
    german: Optional[str] = None
    polish: Optional[str] = None
    english: Optional[str] = None
    french: Optional[str] = None
    italian: Optional[str] = None
    portuguese: Optional[str] = None
    dutch: Optional[str] = None
    russian: Optional[str] = None
    glossary_id: Optional[int] = None
    # Slot-based entries with language_config
    slot1: Optional[str] = None
    slot2: Optional[str] = None
    slot3: Optional[str] = None
    slot4: Optional[str] = None
    slot5: Optional[str] = None
    slot6: Optional[str] = None
    language_config: Optional[list] = None  # e.g., ["german", "spanish", "french", "english"]


class CreateGlossaryRequest(BaseModel):
    name: str


class ApiKeyRequest(BaseModel):
    service: str
    api_key: str


# ==================== Auth Routes ====================


@app.get("/", response_class=HTMLResponse)
async def root(request: Request, db: Session = Depends(get_db)):
    user = await get_current_user(request, db)
    if user:
        return RedirectResponse(url="/translator", status_code=303)
    return RedirectResponse(url="/login", status_code=303)


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, db: Session = Depends(get_db)):
    user = await get_current_user(request, db)
    if user:
        return RedirectResponse(url="/translator", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login", response_class=HTMLResponse)
async def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    user = authenticate_user(db, username, password)
    if not user:
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Invalid username or password"},
        )

    # Email verification is no longer required for login
    # Users can log in and use the app, but unverified email is shown as a hint

    access_token = create_access_token(
        data={"sub": user.username},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
    )

    response = RedirectResponse(url="/translator", status_code=303)
    response.set_cookie(
        key="access_token",
        value=access_token,
        httponly=True,
        max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        samesite="lax",
    )
    return response


@app.get("/register", response_class=HTMLResponse)
async def register_page(request: Request, db: Session = Depends(get_db)):
    user = await get_current_user(request, db)
    if user:
        return RedirectResponse(url="/translator", status_code=303)
    return templates.TemplateResponse("register.html", {"request": request})


@app.post("/register", response_class=HTMLResponse)
async def register(
    request: Request,
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    password_confirm: str = Form(...),
    db: Session = Depends(get_db),
):
    # Validate input
    if len(username) < 3:
        return templates.TemplateResponse(
            "register.html",
            {"request": request, "error": "Username must be at least 3 characters"},
        )

    if len(password) < 6:
        return templates.TemplateResponse(
            "register.html",
            {"request": request, "error": "Password must be at least 6 characters"},
        )

    if password != password_confirm:
        return templates.TemplateResponse(
            "register.html",
            {"request": request, "error": "Passwords do not match"},
        )

    if not email or "@" not in email:
        return templates.TemplateResponse(
            "register.html",
            {"request": request, "error": "Bitte gib eine gültige E-Mail-Adresse ein"},
        )

    # Check if user already exists
    existing_user = get_user_by_username(db, username)
    if existing_user:
        return templates.TemplateResponse(
            "register.html",
            {"request": request, "error": "Username already exists"},
        )

    # Check if email already exists
    existing_email = get_user_by_email(db, email)
    if existing_email:
        return templates.TemplateResponse(
            "register.html",
            {"request": request, "error": "Diese E-Mail-Adresse ist bereits registriert"},
        )

    # Create user
    user = create_user(db, username, password, email)

    # First user (admin) - skip email verification, log in directly
    is_admin = user.id == 1 or user.username == "admin"
    if is_admin:
        user.email_verified = True
        db.commit()
        access_token = create_access_token(
            data={"sub": user.username},
            expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
        )
        response = RedirectResponse(url="/translator", status_code=303)
        response.set_cookie(key="access_token", value=access_token, httponly=True, samesite="lax")
        return response

    # Send verification email in background (non-blocking)
    token = create_verification_token(email)
    base_url = get_base_url(request)
    send_verification_email(email, token, base_url)

    # Log user in directly after registration
    access_token = create_access_token(
        data={"sub": user.username},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
    )
    response = RedirectResponse(url="/translator", status_code=303)
    response.set_cookie(key="access_token", value=access_token, httponly=True, samesite="lax")
    return response


@app.get("/verify-pending", response_class=HTMLResponse)
async def verify_pending_page(request: Request, username: str = Query(None)):
    return templates.TemplateResponse(
        "verify_pending.html",
        {"request": request, "username": username}
    )


@app.post("/resend-verification", response_class=HTMLResponse)
async def resend_verification(
    request: Request,
    username: str = Form(...),
    db: Session = Depends(get_db),
):
    user = get_user_by_username(db, username)
    if user and user.email and not user.email_verified:
        token = create_verification_token(user.email)
        base_url = get_base_url(request)
        send_verification_email(user.email, token, base_url)

    return templates.TemplateResponse(
        "verify_pending.html",
        {"request": request, "username": username, "message": "Bestätigungs-E-Mail wurde erneut gesendet."}
    )


@app.get("/verify-email", response_class=HTMLResponse)
async def verify_email(
    request: Request,
    token: str = Query(...),
    db: Session = Depends(get_db),
):
    email = verify_email_token(token)
    if not email:
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Ungültiger oder abgelaufener Bestätigungslink."}
        )

    user = get_user_by_email(db, email)
    if not user:
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Benutzer nicht gefunden."}
        )

    user.email_verified = True
    db.commit()

    # Auto-login after verification
    access_token = create_access_token(
        data={"sub": user.username},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
    )

    response = RedirectResponse(url="/translator", status_code=303)
    response.set_cookie(
        key="access_token",
        value=access_token,
        httponly=True,
        max_age=ACCESS_TOKEN_EXPIRE_MINUTES * 60,
        samesite="lax",
    )
    return response


@app.get("/change-password", response_class=HTMLResponse)
async def change_password_page(request: Request, db: Session = Depends(get_db)):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(
        "change_password.html", {"request": request, "user": user}
    )


@app.post("/change-password", response_class=HTMLResponse)
async def change_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    new_password_confirm: str = Form(...),
    db: Session = Depends(get_db),
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    if not verify_password(current_password, user.password_hash):
        return templates.TemplateResponse(
            "change_password.html",
            {"request": request, "user": user, "error": "Aktuelles Passwort ist falsch"}
        )

    if len(new_password) < 6:
        return templates.TemplateResponse(
            "change_password.html",
            {"request": request, "user": user, "error": "Neues Passwort muss mindestens 6 Zeichen lang sein"}
        )

    if new_password != new_password_confirm:
        return templates.TemplateResponse(
            "change_password.html",
            {"request": request, "user": user, "error": "Passwörter stimmen nicht überein"}
        )

    user.password_hash = get_password_hash(new_password)
    db.commit()

    return templates.TemplateResponse(
        "change_password.html",
        {"request": request, "user": user, "success": "Passwort wurde erfolgreich geändert"}
    )


@app.get("/forgot-password", response_class=HTMLResponse)
async def forgot_password_page(request: Request):
    return templates.TemplateResponse("forgot_password.html", {"request": request})


@app.post("/forgot-password", response_class=HTMLResponse)
async def forgot_password(
    request: Request,
    email: str = Form(...),
    db: Session = Depends(get_db),
):
    user = get_user_by_email(db, email)
    if user:
        token = create_password_reset_token(email)
        base_url = get_base_url(request)
        send_password_reset_email(email, token, base_url)

    # Always show success to prevent email enumeration
    return templates.TemplateResponse(
        "forgot_password.html",
        {"request": request, "message": "Falls ein Konto mit dieser E-Mail existiert, wurde ein Reset-Link gesendet."}
    )


@app.get("/reset-password", response_class=HTMLResponse)
async def reset_password_page(request: Request, token: str = Query(...)):
    return templates.TemplateResponse(
        "reset_password.html",
        {"request": request, "token": token}
    )


@app.post("/reset-password", response_class=HTMLResponse)
async def reset_password(
    request: Request,
    token: str = Form(...),
    new_password: str = Form(...),
    new_password_confirm: str = Form(...),
    db: Session = Depends(get_db),
):
    email = verify_password_reset_token(token)
    if not email:
        return templates.TemplateResponse(
            "reset_password.html",
            {"request": request, "token": token, "error": "Ungültiger oder abgelaufener Reset-Link."}
        )

    if len(new_password) < 6:
        return templates.TemplateResponse(
            "reset_password.html",
            {"request": request, "token": token, "error": "Passwort muss mindestens 6 Zeichen lang sein"}
        )

    if new_password != new_password_confirm:
        return templates.TemplateResponse(
            "reset_password.html",
            {"request": request, "token": token, "error": "Passwörter stimmen nicht überein"}
        )

    user = get_user_by_email(db, email)
    if not user:
        return templates.TemplateResponse(
            "reset_password.html",
            {"request": request, "token": token, "error": "Benutzer nicht gefunden."}
        )

    user.password_hash = get_password_hash(new_password)
    db.commit()

    return templates.TemplateResponse(
        "login.html",
        {"request": request, "error": "Passwort wurde zurückgesetzt. Du kannst dich jetzt einloggen."}
    )


@app.get("/logout")
async def logout():
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie("access_token")
    return response


# ==================== User API Key Routes ====================


@app.get("/user/api-keys")
async def get_user_api_keys(request: Request, db: Session = Depends(get_db)):
    """Get masked API keys for the current user."""
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    keys = db.query(UserApiKey).filter(UserApiKey.user_id == user.id).all()
    result = {}
    for key in keys:
        try:
            plain = decrypt_api_key(key.api_key)
            # Mask: show first 3 and last 3 chars
            if len(plain) > 8:
                masked = plain[:3] + "***" + plain[-3:]
            else:
                masked = "***"
            result[key.service] = masked
        except Exception:
            result[key.service] = "***"

    trial_days = get_trial_days_remaining(user)

    return {
        "keys": result,
        "trial_days_remaining": trial_days
    }


@app.post("/user/api-keys")
async def save_user_api_key(
    request: Request,
    key_request: ApiKeyRequest,
    db: Session = Depends(get_db),
):
    """Save or update an API key for the current user."""
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    valid_services = ["deepl", "pons", "google"]
    if key_request.service not in valid_services:
        raise HTTPException(status_code=400, detail="Invalid service")

    if not key_request.api_key.strip():
        raise HTTPException(status_code=400, detail="API key cannot be empty")

    encrypted = encrypt_api_key(key_request.api_key.strip())

    # Upsert
    existing = db.query(UserApiKey).filter(
        UserApiKey.user_id == user.id,
        UserApiKey.service == key_request.service
    ).first()

    if existing:
        existing.api_key = encrypted
        existing.created_at = datetime.utcnow()
    else:
        new_key = UserApiKey(
            user_id=user.id,
            service=key_request.service,
            api_key=encrypted,
        )
        db.add(new_key)

    db.commit()
    return {"success": True}


@app.delete("/user/api-keys/{service}")
async def delete_user_api_key(
    request: Request,
    service: str,
    db: Session = Depends(get_db),
):
    """Delete a user's API key for a service."""
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    key = db.query(UserApiKey).filter(
        UserApiKey.user_id == user.id,
        UserApiKey.service == service
    ).first()

    if key:
        db.delete(key)
        db.commit()

    return {"success": True}


# ==================== User Settings Routes ====================

@app.get("/user/settings")
async def get_user_settings(request: Request, db: Session = Depends(get_db)):
    """Get user's language settings from DB."""
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return {"language_config": user.language_config or ""}


@app.post("/user/settings")
async def save_user_settings(request: Request, db: Session = Depends(get_db)):
    """Save user's language settings to DB."""
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    body = await request.json()
    user.language_config = body.get("language_config", "")
    db.commit()
    return {"success": True}


@app.post("/admin/verify-user/{username}")
async def admin_verify_user(request: Request, username: str, db: Session = Depends(get_db)):
    """Admin can manually verify a user's email."""
    admin = await get_current_user(request, db)
    if not admin or (admin.id != 1 and admin.username != "admin"):
        raise HTTPException(status_code=403, detail="Admin only")
    target = get_user_by_username(db, username)
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    target.email_verified = True
    db.commit()
    return {"success": True, "message": f"User {username} verified"}


# ==================== Translator Routes ====================


@app.get("/translator", response_class=HTMLResponse)
async def translator_page(request: Request, db: Session = Depends(get_db)):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(
        "translator.html", {"request": request, "user": user}
    )


@app.post("/translate")
async def translate(
    request: Request,
    translate_request: TranslateRequest,
    db: Session = Depends(get_db),
):
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    text = translate_request.text.strip()
    if not text:
        raise HTTPException(status_code=400, detail="Text is required")

    translations = translate_to_all_languages(
        text,
        translate_request.source_language,
        translate_request.target_languages,
        translate_request.enabled_services,
        translate_request.explanation_services,
        user_id=user.id,
        db=db,
    )
    return translations


# ==================== Glossary Routes ====================


def get_or_create_default_glossary(db: Session, user_id: int) -> Glossary:
    """Get user's default glossary or create one if it doesn't exist."""
    default = db.query(Glossary).filter(
        Glossary.user_id == user_id,
        Glossary.is_default == True
    ).first()

    if not default:
        default = Glossary(
            user_id=user_id,
            name="Hauptglossar",
            is_default=True
        )
        db.add(default)
        db.commit()
        db.refresh(default)

    return default


@app.get("/glossaries")
async def list_glossaries(request: Request, db: Session = Depends(get_db)):
    """List all glossaries for the current user."""
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    # Ensure default glossary exists
    get_or_create_default_glossary(db, user.id)

    glossaries = db.query(Glossary).filter(
        Glossary.user_id == user.id
    ).order_by(Glossary.is_default.desc(), Glossary.name).all()

    return [
        {
            "id": g.id,
            "name": g.name,
            "is_default": g.is_default,
            "entry_count": len(g.entries)
        }
        for g in glossaries
    ]


@app.post("/glossaries")
async def create_glossary(
    request: Request,
    glossary_request: CreateGlossaryRequest,
    db: Session = Depends(get_db),
):
    """Create a new glossary."""
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    name = glossary_request.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name is required")

    # Check if glossary with same name exists
    existing = db.query(Glossary).filter(
        Glossary.user_id == user.id,
        Glossary.name == name
    ).first()

    if existing:
        raise HTTPException(status_code=400, detail="Glossary with this name already exists")

    glossary = Glossary(
        user_id=user.id,
        name=name,
        is_default=False
    )
    db.add(glossary)
    db.commit()
    db.refresh(glossary)

    return {
        "id": glossary.id,
        "name": glossary.name,
        "is_default": glossary.is_default,
        "entry_count": 0
    }


@app.post("/glossary/save")
async def save_to_glossary(
    request: Request,
    glossary_request: SaveGlossaryRequest,
    db: Session = Depends(get_db),
):
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    # Get glossary (use specified or default)
    if glossary_request.glossary_id:
        glossary = db.query(Glossary).filter(
            Glossary.id == glossary_request.glossary_id,
            Glossary.user_id == user.id
        ).first()
        if not glossary:
            raise HTTPException(status_code=404, detail="Glossary not found")
    else:
        glossary = get_or_create_default_glossary(db, user.id)

    ALL_LANG_COLS = ["spanish", "german", "polish", "english", "french", "italian", "portuguese", "dutch", "russian"]

    # Handle slot-based format with language_config
    if glossary_request.language_config and glossary_request.slot1 is not None:
        lang_to_slot = {}
        slots = [glossary_request.slot1, glossary_request.slot2,
                 glossary_request.slot3, glossary_request.slot4,
                 glossary_request.slot5, glossary_request.slot6]
        for i, lang in enumerate(glossary_request.language_config):
            if i < len(slots) and slots[i] is not None:
                lang_to_slot[lang] = slots[i]

        entry = GlossaryEntry(
            user_id=user.id,
            glossary_id=glossary.id,
            **{col: lang_to_slot.get(col, "") for col in ALL_LANG_COLS}
        )
    else:
        entry = GlossaryEntry(
            user_id=user.id,
            glossary_id=glossary.id,
            **{col: getattr(glossary_request, col) or "" for col in ALL_LANG_COLS}
        )
    db.add(entry)
    db.commit()

    return {"success": True, "message": f"Entry saved to {glossary.name}"}


@app.get("/glossary/recent")
async def get_recent_entries(
    request: Request,
    glossary_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Get the 4 most recent entries for a glossary."""
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    # Get specific glossary or default
    if glossary_id:
        glossary = db.query(Glossary).filter(
            Glossary.id == glossary_id,
            Glossary.user_id == user.id
        ).first()
        if not glossary:
            raise HTTPException(status_code=404, detail="Glossary not found")
    else:
        glossary = get_or_create_default_glossary(db, user.id)

    # Get 4 most recent entries
    entries = (
        db.query(GlossaryEntry)
        .filter(GlossaryEntry.glossary_id == glossary.id)
        .order_by(GlossaryEntry.created_at.desc())
        .limit(4)
        .all()
    )

    ALL_LANG_COLS = ["spanish", "german", "polish", "english", "french", "italian", "portuguese", "dutch", "russian"]
    return [
        {
            **{col: getattr(e, col) for col in ALL_LANG_COLS},
            "entries": {col: getattr(e, col) for col in ALL_LANG_COLS}
        }
        for e in entries
    ]


@app.get("/glossary/export")
async def export_glossary(
    request: Request,
    glossary_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    # Get specific glossary or default
    if glossary_id:
        glossary = db.query(Glossary).filter(
            Glossary.id == glossary_id,
            Glossary.user_id == user.id
        ).first()
        if not glossary:
            raise HTTPException(status_code=404, detail="Glossary not found")
    else:
        glossary = get_or_create_default_glossary(db, user.id)

    # Get entries for the glossary
    entries = (
        db.query(GlossaryEntry)
        .filter(GlossaryEntry.glossary_id == glossary.id)
        .order_by(GlossaryEntry.created_at.desc())
        .all()
    )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = glossary.name[:31]  # Excel sheet names max 31 chars

    # Determine which language columns have data
    ALL_LANG_COLS = ["spanish", "german", "polish", "english", "french", "italian", "portuguese", "dutch", "russian"]
    active_cols = []
    for col in ALL_LANG_COLS:
        if any(getattr(e, col) for e in entries):
            active_cols.append(col)
    if not active_cols:
        active_cols = ALL_LANG_COLS[:4]

    # Header row
    headers = [col.capitalize() for col in active_cols] + ["Created At"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)

    # Data rows
    for row, entry in enumerate(entries, 2):
        for col_idx, col in enumerate(active_cols, 1):
            ws.cell(row=row, column=col_idx, value=getattr(entry, col))
        ws.cell(row=row, column=len(active_cols) + 1, value=entry.created_at.strftime("%Y-%m-%d %H:%M"))

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{glossary.name}_{user.username}.xlsx".replace(" ", "_")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ==================== Glossary List Routes ====================


@app.get("/glossary-list", response_class=HTMLResponse)
async def glossary_list_page(request: Request, db: Session = Depends(get_db)):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(
        "glossary_list.html", {"request": request, "user": user}
    )


@app.get("/glossary/entries")
async def get_glossary_entries(
    request: Request,
    glossary_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """Get all entries for a glossary."""
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    if glossary_id:
        glossary = db.query(Glossary).filter(
            Glossary.id == glossary_id,
            Glossary.user_id == user.id
        ).first()
        if not glossary:
            raise HTTPException(status_code=404, detail="Glossary not found")
    else:
        glossary = get_or_create_default_glossary(db, user.id)

    entries = (
        db.query(GlossaryEntry)
        .filter(GlossaryEntry.glossary_id == glossary.id)
        .order_by(GlossaryEntry.created_at.desc())
        .all()
    )

    ALL_LANG_COLS = ["spanish", "german", "polish", "english", "french", "italian", "portuguese", "dutch", "russian"]
    return [
        {
            "id": e.id,
            **{col: getattr(e, col) for col in ALL_LANG_COLS},
            "total_learning_rate": e.total_learning_rate or 0,
            "created_at": e.created_at.isoformat() if e.created_at else None,
        }
        for e in entries
    ]


@app.delete("/glossary/entry/{entry_id}")
async def delete_glossary_entry(
    request: Request,
    entry_id: int,
    db: Session = Depends(get_db),
):
    """Delete a glossary entry."""
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    entry = db.query(GlossaryEntry).filter(
        GlossaryEntry.id == entry_id,
        GlossaryEntry.user_id == user.id,
    ).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    db.delete(entry)
    db.commit()
    return {"success": True}


# ==================== Vocab Test Routes ====================


@app.get("/vocab-test", response_class=HTMLResponse)
async def vocab_test_page(request: Request, db: Session = Depends(get_db)):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse(
        "vocab_test.html", {"request": request, "user": user}
    )


class VocabStartRequest(BaseModel):
    glossary_id: Optional[int] = None
    days: int = 1
    learn_limit: int = 10


@app.post("/vocab-test/start")
async def vocab_test_start(
    request: Request,
    start_request: VocabStartRequest,
    db: Session = Depends(get_db),
):
    """Reset learning_rate to 0 for all entries in scope, return count."""
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    if start_request.glossary_id:
        glossary = db.query(Glossary).filter(
            Glossary.id == start_request.glossary_id, Glossary.user_id == user.id
        ).first()
        if not glossary:
            raise HTTPException(status_code=404, detail="Glossary not found")
    else:
        glossary = get_or_create_default_glossary(db, user.id)

    since = datetime.utcnow() - timedelta(days=start_request.days)

    # Reset learning_rate for entries in scope
    count = (
        db.query(GlossaryEntry)
        .filter(
            GlossaryEntry.glossary_id == glossary.id,
            GlossaryEntry.created_at >= since,
            sa_func.coalesce(GlossaryEntry.total_learning_rate, 0) < start_request.learn_limit,
        )
        .update({"learning_rate": 0}, synchronize_session="fetch")
    )
    db.commit()
    return {"success": True, "reset_count": count}


ALL_LANG_COLS = ["spanish", "german", "polish", "english", "french", "italian", "portuguese", "dutch", "russian"]


@app.get("/vocab-test/entries")
async def vocab_test_entries(
    request: Request,
    question_lang: str,
    answer_langs: str = "",
    glossary_id: Optional[int] = None,
    days: int = 1,
    max_rate: int = 3,
    learn_limit: int = 10,
    db: Session = Depends(get_db),
):
    """Return all matching entries for the client to shuffle and iterate."""
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    if glossary_id:
        glossary = db.query(Glossary).filter(
            Glossary.id == glossary_id, Glossary.user_id == user.id
        ).first()
        if not glossary:
            raise HTTPException(status_code=404, detail="Glossary not found")
    else:
        glossary = get_or_create_default_glossary(db, user.id)

    since = datetime.utcnow() - timedelta(days=days)

    question_col = getattr(GlossaryEntry, question_lang, None)
    if question_col is None:
        raise HTTPException(status_code=400, detail="Invalid language")

    entries = (
        db.query(GlossaryEntry)
        .filter(
            GlossaryEntry.glossary_id == glossary.id,
            GlossaryEntry.created_at >= since,
            sa_func.coalesce(GlossaryEntry.learning_rate, 0) < max_rate,
            sa_func.coalesce(GlossaryEntry.total_learning_rate, 0) < learn_limit,
            question_col.isnot(None),
            question_col != "",
        )
        .all()
    )

    a_langs = [l for l in answer_langs.split(",") if l and l != question_lang] if answer_langs else []

    result = []
    for entry in entries:
        answers = {}
        for lang in a_langs:
            val = getattr(entry, lang, None)
            if val:
                answers[lang] = val
        result.append({
            "id": entry.id,
            "question": getattr(entry, question_lang),
            "answers": answers,
            "learning_rate": entry.learning_rate or 0,
            "total_learning_rate": entry.total_learning_rate or 0,
        })

    return {"entries": result, "total": len(result)}


class VocabAnswerRequest(BaseModel):
    entry_id: int
    correct: bool


@app.post("/vocab-test/answer")
async def vocab_test_answer(
    request: Request,
    answer_request: VocabAnswerRequest,
    db: Session = Depends(get_db),
):
    user = await get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")

    entry = db.query(GlossaryEntry).filter(
        GlossaryEntry.id == answer_request.entry_id,
        GlossaryEntry.user_id == user.id,
    ).first()

    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    if answer_request.correct:
        entry.learning_rate = (entry.learning_rate or 0) + 1
        entry.total_learning_rate = (entry.total_learning_rate or 0) + 1
        db.commit()

    return {
        "success": True,
        "learning_rate": entry.learning_rate or 0,
        "total_learning_rate": entry.total_learning_rate or 0,
    }


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)
